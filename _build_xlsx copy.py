"""Build the two companion xlsx files for the Magic Scarf rebuild."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ===== shared styling =====
HDR_FILL = PatternFill('solid', start_color='8B2942')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
ZEBRA_FILL = PatternFill('solid', start_color='FAF6F1')
SUB_FILL = PatternFill('solid', start_color='F4ECE6')
SUB_FONT = Font(name='Arial', bold=True, color='1F1A1D', size=11)
BODY_FONT = Font(name='Arial', size=10, color='1F1A1D')
THIN = Side(border_style='thin', color='E6DDDB')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_header(sheet, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = sheet.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    sheet.row_dimensions[row].height = 32


def style_body(sheet, start_row, end_row, n_cols, zebra=True):
    for r in range(start_row, end_row + 1):
        is_zebra = zebra and (r - start_row) % 2 == 1
        for c in range(1, n_cols + 1):
            cell = sheet.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_TOP
            cell.border = BORDER
            if is_zebra:
                cell.fill = ZEBRA_FILL


def set_widths(sheet, widths):
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 1. component-image-sheet.xlsx
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = 'Components & Images'

ws['A1'] = 'Magic Scarf — Component-Wise Image & Content Sheet'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='8B2942')
ws.merge_cells('A1:H1')

ws['A2'] = 'Each row maps a UI component to its image / content requirement at desktop, tablet, and mobile.  Use this as the production asset checklist.'
ws['A2'].font = Font(name='Arial', italic=True, size=10, color='4A4248')
ws['A2'].alignment = Alignment(wrap_text=True)
ws.merge_cells('A2:H2')

headers = ['#', 'Page / View', 'Component', 'Image (W × H px)  Desktop', 'Image (W × H px)  Tablet', 'Image (W × H px)  Mobile', 'Format', 'Content / Notes']
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

rows = [
    # Header chrome
    ('Global', 'Logo (header)', '180 × 56', '160 × 50', '120 × 40', 'PNG/SVG (transparent)', 'Use audited logo.png from CDN. Customer to supply higher-res or SVG.'),
    ('Global', 'Favicon / app-icon', '144 × 144', '144 × 144', '144 × 144', 'PNG', 'Audited from /assets/icons/ms-icon-144x144.png — keep as-is.'),
    ('Global', 'Footer logo (dark bg)', '180 × 56', '160 × 50', '120 × 40', 'PNG/SVG (white version)', 'White or knockout version of logo for dark footer.'),

    # Home — hero
    ('Home', 'Hero main image', '720 × 900', '600 × 750', '375 × 470', 'JPG (≤300 KB)', 'Primary lifestyle / product hero. Audit URL: rotate_1291.main.jpg. Replace with new shot if available.'),
    ('Home', 'Hero thumbnails (×6)', '128 × 128', '110 × 110', '64 × 64', 'JPG (≤40 KB each)', 'Rotators 1281 / 1282 / 1285 / 1290 / 1291 / 1292. 6 thumbs total. Manual select — no autoplay.'),

    # Home — seasonal rail
    ('Home', 'Seasonal category tile (×6)', '320 × 320', '300 × 300', '180 × 180', 'JPG square', 'category_ad_<id>.jpg pattern. 6 of: Oblong, Infinity, Shawls/Wraps, Hats, Shanghai, Lightweight.'),

    # Home — catalog grid
    ('Home', 'Catalog tile (×17)', '480 × 384', '420 × 336', '320 × 256', 'JPG 5:4', 'category_<id>.jpg pattern (599, 352, 941, 957, 919, 853, 340, 826, 335, 914, 662, 496, 741, 297, 633, 865, 612).'),

    # Home — story band
    ('Home', 'Story band image', '900 × 675', '760 × 570', '375 × 280', 'JPG 4:3', 'Founder studio shot, showroom photo, or aspirational lifestyle. Customer to supply.'),

    # Home — show schedule
    ('Home', 'Show card date pill (×4)', 'inline SVG', 'inline SVG', 'inline SVG', 'CSS only', 'No image needed — date pill rendered in CSS with month + day text.'),

    # Home — Faire band
    ('Home', 'Faire logo / wordmark', '120 × 32', '110 × 30', '80 × 22', 'PNG (transparent)', 'Faire brand mark. Optional — currently uses text only.'),

    # Home — signup / press
    ('Home', "Press 'As Seen In' logos (×6)", '160 × 60', '140 × 50', '100 × 36', 'PNG monochrome', 'Optional. Currently not built — add when customer supplies press logos.'),

    # PLP
    ('PLP', 'Subcategory pill image (optional)', '40 × 40', '40 × 40', '40 × 40', 'JPG circle', 'Optional thumb on each subcategory pill. Currently text-only.'),
    ('PLP', 'Product card thumbnail', '480 × 600', '420 × 525', '320 × 400', 'JPG 4:5', '/i/p/<pid>/<n>/300x300.jpg pattern from audit. Card aspect ratio 4:5.'),
    ('PLP', 'Filter color swatch (×8)', '32 × 32', '32 × 32', '32 × 32', 'CSS color', 'No image — CSS chip with border.'),

    # PDP
    ('PDP', 'Main gallery image', '900 × 1125', '760 × 950', '375 × 470', 'JPG 4:5 (≤300 KB)', 'main_image_<pid>.jpg pattern from audit. Sticky on desktop.'),
    ('PDP', 'Thumb strip (×4)', '160 × 192', '140 × 168', '84 × 96', 'JPG ~5:6', '/i/p/<pid>/<n>/300x300.jpg with n ∈ {3,8,13,17}.'),
    ('PDP', 'Variant tile (×N — up to 47)', '120 × 120', '110 × 110', '90 × 90', 'JPG square or CSS', '/i/o/<pid>/<color_id>/<n>/600x600.jpg if available, else CSS color chip with name overlay. Magic Scarf has 47 colors for SKU 8672.'),
    ('PDP', 'Spec sheet PDF (download)', 'PDF', 'PDF', 'PDF', 'PDF', "Audited as 'Printable Swatch Page'. Customer to supply or auto-generate."),
    ('PDP', 'Product video (YouTube)', '1280 × 720 (player)', '1024 × 576', '375 × 211', 'YouTube embed', 'Audited from PDP — Magic Scarf YouTube channel UCXLskc4sAfuppKmS-oJZ1Rw.'),

    # Cart
    ('Cart', 'Cart row product thumb', '192 × 192', '160 × 160', '96 × 96', 'JPG square', 'main_image_<pid>.jpg cropped square.'),

    # Account
    ('Account', 'Avatar (initials or photo)', '104 × 104', '88 × 88', '88 × 88', 'JPG circle or initials', 'Initials default; user can upload photo later.'),

    # Modals
    ('Login modal', 'Promo panel background', '440 × 600', 'n/a (stacks)', 'n/a', 'CSS gradient', 'No image — Magic Burgundy gradient with radial accent.'),

    # Footer
    ('Global', 'Payment provider pills (×4)', '60 × 24', '50 × 20', '40 × 18', 'CSS only', 'VISA / MC / AMEX / NET-30 — text pills, no image needed.'),
    ('Global', 'Social platform icons (×2)', '24 × 24', '24 × 24', '20 × 20', 'inline SVG', 'Facebook + YouTube only per audit.'),

    # Hero rotators full set
    ('Home', 'Hero rotator set (full pool)', '1280 × 800', '900 × 600', '375 × 470', 'JPG (≤350 KB each)', "Pool of 7 rotators audited: 1266 (Faire), 1281, 1282, 1285, 1290, 1291, 1292. Customer can refresh per season."),
]

for i, row in enumerate(rows):
    n = i + 5
    ws.cell(row=n, column=1, value=i + 1)
    for j, val in enumerate(row, start=2):
        ws.cell(row=n, column=j, value=val)

style_body(ws, 5, 4 + len(rows), len(headers), zebra=True)
set_widths(ws, [5, 12, 32, 22, 22, 22, 18, 60])
ws.row_dimensions[1].height = 28

# Sheet 2 — Asset URL inventory
ws2 = wb.create_sheet('Asset URL Inventory')
ws2['A1'] = 'Magic Scarf — Audited Asset URL Inventory'
ws2['A1'].font = Font(name='Arial', bold=True, size=16, color='8B2942')
ws2.merge_cells('A1:D1')
ws2['A2'] = 'Real URLs harvested from the live site. Use these for v1; customer can swap to new CDN later without code changes.'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='4A4248')
ws2.merge_cells('A2:D2')

headers2 = ['Type', 'Pattern', 'Example', 'Notes']
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(headers2))

asset_rows = [
    ('Logo', 'fixed URL', 'https://www.magicscarf.com/stores/eol/assets/images/logo.png', 'Single source of truth for logo.'),
    ('Favicon', 'fixed URL', 'https://www.magicscarf.com/stores/eol/assets/icons/ms-icon-144x144.png', '144×144 PNG.'),
    ('Category tile (catalog)', '/stores/common_files/categories/category_<ID>.jpg', 'category_599.jpg / category_352.jpg / category_914.jpg', '17 IDs: 599, 352, 941, 957, 919, 853, 340, 826, 335, 914, 662, 496, 741, 297, 633, 865, 612.'),
    ('Category tile (seasonal rail)', '/stores/common_files/categories/category_ad_<ID>.jpg', 'category_ad_940.jpg / category_ad_939.jpg', '13 seasonal: 940, 939, 941, 957, 434, 533, 340, 624, 496, 627, 741, 612, 947.'),
    ('Hero rotator', '/rotators/rotate_<RID>.main.jpg (S3)', 'https://com-magicscarf-web.s3.amazonaws.com/rotators/rotate_1291.main.jpg', 'Audited rotator pool: 1266 (Faire), 1281, 1282, 1285, 1290, 1291, 1292.'),
    ('Product main image', '/stores/common_files/products/main_image_<PID>.jpg', 'main_image_3103.jpg / main_image_2997.jpg', 'PID is the product internal ID (4-digit).'),
    ('Product card thumb', '/i/p/<PID>/<N>/300x300.jpg', '/i/p/3103/3/300x300.jpg', 'N is the image variant index (1–60+ for products with many photos).'),
    ('Variant image', '/i/o/<PID>/<COLOR_ID>/<N>/600x600.jpg', '/i/o/3103/79355/1/600x600.jpg', 'COLOR_ID is internal — for 8672 Cashmere Feel Poncho there are 47 variant IDs in range 79343–79404.'),
    ('Product detail (pretty)', '/stores/eol/p-<PID>-<slug>.html', '/stores/eol/p-3103-poncho-solid-cashmere-feel-8672.html', 'Customer SEO URL pattern. WizCommerce equivalent: /products/<slug>.'),
    ('Category (pretty)', '/stores/eol/c-<ID>-<slug>.html', '/stores/eol/c-914-ponchos.html', 'Used as a 301 redirect source post-migration.'),
    ('YouTube channel', 'fixed URL', 'https://www.youtube.com/channel/UCXLskc4sAfuppKmS-oJZ1Rw/', "Magic Scarf 'Product Videos' link in audited footer."),
    ('Faire storefront', 'fixed URL', 'https://www.faire.com/direct/magicscarf', 'External marketplace mirror.'),
    ('Facebook', 'fixed URL', 'https://www.facebook.com/magicscarf', 'Only social platforms audited.'),
]
for i, row in enumerate(asset_rows):
    n = i + 5
    for j, val in enumerate(row, start=1):
        ws2.cell(row=n, column=j, value=val)
style_body(ws2, 5, 4 + len(asset_rows), len(headers2), zebra=True)
set_widths(ws2, [22, 48, 60, 60])

wb.save('/sessions/lucid-quirky-lamport/mnt/outputs/magic-scarf-site/component-image-sheet.xlsx')
print('Wrote component-image-sheet.xlsx')

# ============================================================
# 2. client-approval-tracker.xlsx
# ============================================================
wb2 = Workbook()
ws_a = wb2.active
ws_a.title = 'Approval Tracker'

ws_a['A1'] = 'Magic Scarf — Client Approval Tracker'
ws_a['A1'].font = Font(name='Arial', bold=True, size=16, color='8B2942')
ws_a.merge_cells('A1:I1')

ws_a['A2'] = 'One row per page × component. Customer signs each off via column status. Owner column tracks who is responsible. Use comments column for change requests.'
ws_a['A2'].font = Font(name='Arial', italic=True, size=10, color='4A4248')
ws_a.merge_cells('A2:I2')

ws_a['A3'] = "Color key:  Green = approved · Yellow = customer reviewing · Red = blocked / change requested · Blue = WizCommerce internal"
ws_a['A3'].font = Font(name='Arial', italic=True, size=9, color='4A4248')
ws_a.merge_cells('A3:I3')

# Approval-summary block (uses formulas to count statuses)
ws_a['A5'] = 'Total rows'
ws_a['A6'] = 'Approved'
ws_a['A7'] = 'In review'
ws_a['A8'] = 'Blocked'
ws_a['A9'] = 'Approval %'
for c in ['A5','A6','A7','A8','A9']:
    ws_a[c].font = Font(name='Arial', bold=True, size=10)

# Will fill formulas after data is laid out — placeholder cells
ws_a['B5'] = '=COUNTA(B14:B200)'
ws_a['B6'] = '=COUNTIF(F14:F200,"Approved")'
ws_a['B7'] = '=COUNTIF(F14:F200,"In review")'
ws_a['B8'] = '=COUNTIF(F14:F200,"Blocked")'
ws_a['B9'] = '=IFERROR(B6/B5,0)'
ws_a['B9'].number_format = '0.0%'

# Header row (row 13)
hdr_row = 13
hdr = ['#', 'Page / View', 'Component / Section', 'Owner (WizComm)', 'Owner (Magic Scarf)', 'Status', 'Sign-off date', 'Customer comments', 'Resolved by']
for i, h in enumerate(hdr, start=1):
    ws_a.cell(row=hdr_row, column=i, value=h)
style_header(ws_a, hdr_row, len(hdr))

# Sign-off rows — every page × major component
items = [
    ('Global', 'Logo + wordmark fallback'),
    ('Global', 'Brand color (Magic Burgundy #8B2942)'),
    ('Global', 'Type pairing (Cormorant + Inter)'),
    ('Global', 'Announcement bar copy + perks'),
    ('Global', 'Sticky header layout'),
    ('Global', 'Search pill UX'),
    ('Global', 'Mobile drawer nav'),
    ('Global', 'Footer 5-column layout + brand block'),
    ('Global', 'Toll-free phone in footer + announce + contact'),
    ('Global', 'Social links (Facebook + YouTube only)'),
    ('Home', 'Hero copy + dual CTAs'),
    ('Home', 'Hero figure (rotator image selection)'),
    ('Home', 'Hero thumb strip (manual carousel)'),
    ('Home', 'Trust line under hero CTAs'),
    ('Home', 'Seasonal Categories rail (6 tiles)'),
    ('Home', 'Magic Scarf Catalog grid (17 tiles)'),
    ('Home', 'Brand Story band copy'),
    ('Home', 'Story stats (20+ years / 12k retailers / 900+ SKUs)'),
    ('Home', 'Trade-show schedule preview (4 cards)'),
    ('Home', 'Faire band copy + CTA'),
    ('Home', 'Wholesale signup band'),
    ('Catalog', 'Catalog grid (17 categories)'),
    ('Catalog', 'Category descriptions'),
    ('PLP — Ponchos', 'Subcategory pill bar'),
    ('PLP — Ponchos', 'Filter sidebar (Subcat / Color / Material / Pattern / Stock)'),
    ('PLP — Ponchos', 'Sort dropdown options'),
    ('PLP — Ponchos', 'Product card layout (SKU + chips + price-locked)'),
    ('PLP — Ponchos', 'Pagination'),
    ('PLP — Scarves & Shawls', 'Audited description copy'),
    ('PLP — Scarves & Shawls', '7 audited subcategory pills'),
    ('PLP — Best Picks', 'Layout + 16-card grid'),
    ('PLP — Sale', 'Discount badge styling'),
    ('PDP — 8672 Cashmere Poncho', 'Gallery (sticky main + 4 thumbs)'),
    ('PDP — 8672 Cashmere Poncho', 'Audited description verbatim'),
    ('PDP — 8672 Cashmere Poncho', '47-color variant grid + named labels'),
    ('PDP — 8672 Cashmere Poncho', 'Sourcing strip (case-pack / MOQ / 24h)'),
    ('PDP — 8672 Cashmere Poncho', 'Login-gated price box'),
    ('PDP — 8672 Cashmere Poncho', 'Trust strip (4 icons)'),
    ('PDP — 8672 Cashmere Poncho', '4 accordions (Details / Specs / Shipping / Care)'),
    ('PDP — 8672 Cashmere Poncho', 'Related Products always-fills-to-4'),
    ('PDP — 7777 Victorian Lace', 'Heritage badge + description'),
    ('PDP — 2997 Magnetic Brooch', 'Accessory PDP layout (different sourcing)'),
    ('Cart', 'Empty state'),
    ('Cart', 'Cart row layout (image + meta + qty stepper + remove)'),
    ('Cart', 'Sticky summary sidebar (login-gated subtotal)'),
    ('Account', 'Sidebar with avatar + tier pill'),
    ('Account', 'KPI tiles (open orders / cart / Net-30)'),
    ('Account', 'Recent orders table'),
    ('Show Schedule', '8-card full grid'),
    ('Show Schedule', 'Real 2026–2027 calendar (CUSTOMER TO SUPPLY)'),
    ('Contact', 'Contact form (7 fields + topic select)'),
    ('Contact', 'Reach card (phone / email / hours / showroom)'),
    ('Login modal', 'Split panel layout'),
    ('Login modal', 'Promo panel copy'),
    ('Register modal', '11 audited fields + tier dropdowns'),
    ('Register modal', 'Apply for Wholesale promo'),
    ('Search modal', 'Placeholder (real search wired post-launch)'),
    ('Mobile (375)', 'Card grid 2-up'),
    ('Mobile (375)', 'Hamburger drawer all 9 items'),
    ('Mobile (375)', 'PDP variant grid scrollability'),
    ('Mobile (375)', 'Sticky header height (64px)'),
    ('Tablet (900)', 'Card grid 3-up'),
    ('Tablet (900)', 'PLP body 1-col below 1100'),
    ('Accessibility', 'Body type ≥17px (older-B2B)'),
    ('Accessibility', 'Buttons ≥52px primary'),
    ('Accessibility', 'No autoplay carousels'),
    ('Accessibility', 'Focus rings visible'),
    ('Accessibility', 'Alt text on every image'),
    ('Accessibility', 'WCAG AA contrast on body'),
    ('Performance', 'No console errors'),
    ('Performance', 'Lazy-loaded images'),
    ('Brand assumption', 'Magic Burgundy #8B2942 confirmed by customer'),
    ('Brand assumption', "47-color hex chips QA'd against fabric"),
    ('Brand assumption', 'Founder photo supplied'),
    ('Brand assumption', 'Showroom address confirmed'),
    ('Brand assumption', 'Founding year 2003 confirmed'),
    ('Backend', 'Real authentication wired'),
    ('Backend', 'Real cart persistence'),
    ('Backend', 'Real search'),
    ('Backend', 'Sort + filter API live'),
    ('Backend', 'Tier-pricing API'),
    ('Migration', 'SKU master CSV uploaded'),
    ('Migration', 'Variant CSV uploaded'),
    ('Migration', 'Image CDN configured'),
    ('Migration', '301 redirects from /stores/eol/ paths'),
]

import datetime
for i, (page, comp) in enumerate(items):
    r = hdr_row + 1 + i
    ws_a.cell(row=r, column=1, value=i + 1)
    ws_a.cell(row=r, column=2, value=page)
    ws_a.cell(row=r, column=3, value=comp)
    ws_a.cell(row=r, column=4, value='Design + Eng')
    ws_a.cell(row=r, column=5, value='—')
    ws_a.cell(row=r, column=6, value='In review')
    ws_a.cell(row=r, column=7, value='')
    ws_a.cell(row=r, column=8, value='')
    ws_a.cell(row=r, column=9, value='')

style_body(ws_a, hdr_row + 1, hdr_row + len(items), len(hdr), zebra=True)
set_widths(ws_a, [5, 22, 48, 18, 18, 16, 14, 36, 16])

# Conditional-style sample for status column — colored fills based on value
from openpyxl.formatting.rule import CellIsRule
GREEN_FILL = PatternFill('solid', start_color='D6EFD3')
YELLOW_FILL = PatternFill('solid', start_color='FFF4D6')
RED_FILL = PatternFill('solid', start_color='F8D7DA')
BLUE_FILL = PatternFill('solid', start_color='D6E6FB')
status_range = f'F{hdr_row+1}:F{hdr_row+len(items)}'
ws_a.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Approved"'], fill=GREEN_FILL, font=Font(name='Arial', bold=True, color='2F5F3A')))
ws_a.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"In review"'], fill=YELLOW_FILL, font=Font(name='Arial', color='8A6D2F')))
ws_a.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Blocked"'], fill=RED_FILL, font=Font(name='Arial', bold=True, color='8B2942')))
ws_a.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"WizComm-internal"'], fill=BLUE_FILL, font=Font(name='Arial', color='1F3FA8')))

# Data validation for status column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Approved,In review,Blocked,WizComm-internal,N/A"', allow_blank=True)
dv.error = 'Pick a status from the dropdown'
dv.errorTitle = 'Invalid status'
dv.add(status_range)
ws_a.add_data_validation(dv)

# Update summary formulas to actual range
last_row = hdr_row + len(items)
ws_a['B5'] = f'=COUNTA(B{hdr_row+1}:B{last_row})'
ws_a['B6'] = f'=COUNTIF(F{hdr_row+1}:F{last_row},"Approved")'
ws_a['B7'] = f'=COUNTIF(F{hdr_row+1}:F{last_row},"In review")'
ws_a['B8'] = f'=COUNTIF(F{hdr_row+1}:F{last_row},"Blocked")'
ws_a['B9'] = '=IFERROR(B6/B5,0)'
ws_a['B9'].number_format = '0.0%'

# Sheet 2 — sign-off page
ws_b = wb2.create_sheet('Sign-off')
ws_b['A1'] = 'Magic Scarf — Final Sign-off'
ws_b['A1'].font = Font(name='Arial', bold=True, size=16, color='8B2942')
ws_b.merge_cells('A1:E1')

signoff_hdr = ['Reviewer', 'Role', 'Decision', 'Date', 'Signature']
for i, h in enumerate(signoff_hdr, start=1):
    ws_b.cell(row=3, column=i, value=h)
style_header(ws_b, 3, len(signoff_hdr))

reviewers = [
    ('', 'Engineering (WizCommerce)', '', '', ''),
    ('', 'Product (WizCommerce)', '', '', ''),
    ('', 'CSM (WizCommerce)', '', '', ''),
    ('', 'Brand owner (Magic Scarf)', '', '', ''),
    ('', 'Operations / Wholesale (Magic Scarf)', '', '', ''),
    ('', 'Approving stakeholder (Magic Scarf)', '', '', ''),
]
for i, row in enumerate(reviewers):
    r = 4 + i
    for j, val in enumerate(row, start=1):
        ws_b.cell(row=r, column=j, value=val)
style_body(ws_b, 4, 3 + len(reviewers), len(signoff_hdr), zebra=True)
set_widths(ws_b, [22, 32, 18, 14, 32])
for r in range(4, 4 + len(reviewers)):
    ws_b.row_dimensions[r].height = 36

wb2.save('/sessions/lucid-quirky-lamport/mnt/outputs/magic-scarf-site/client-approval-tracker.xlsx')
print('Wrote client-approval-tracker.xlsx')
